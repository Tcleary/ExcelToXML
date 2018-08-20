from openpyxl import load_workbook
from lxml import etree

EXCEL_SPREADSHEET = raw_input("EXCEL SPREAD SHEET FILENAME: ")
wb = load_workbook(filename=EXCEL_SPREADSHEET, read_only=True)
ws = wb['Sheet1']

files = []
for row in ws.iter_rows(min_row=3):
		count = 0
		dict = {}
		for cell in row:
			count+= 1
			if count % 16 == 1:
				dict["Filename"] = cell.value
			elif count % 16 == 2:
				dict["Title"] = cell.value
			elif count % 16 == 3:
				dict["Creators"] = cell.value
			elif count % 16 == 4:
				dict["Subjects"] = cell.value
			elif count % 16 == 5:
				dict["Description"] = cell.value
			elif count % 16 == 6:
				dict["Publisher"] = cell.value
			elif count % 16 == 7:
				dict["Contributors"] = cell.value
			elif count % 16 == 8:
				dict["Date"] = cell.value
			elif count % 16 == 9:
				dict["Type"] = cell.value
			elif count % 16 == 10:
				dict["Format"] = cell.value
			elif count % 16 == 11:
				dict["Identifier"] = cell.value
			elif count % 16 == 12:
				dict["Source"] = cell.value
			elif count % 16 == 13:
				dict["Language"] = cell.value
			elif count % 16 == 14:
				dict["Relation"] = cell.value
			elif count % 16 == 15:
				dict["Coverage"] = cell.value
			else:
				dict["Rights"] = cell.value
		files.append(dict)

for f in files:
	current = open(f["Filename"],"w")
	#page = etree.Element("  will use lxml later
	current.write('''<oai_dc:dc xmlns:oai_dc="http://www.openarchives.org/OAI/2.0/oai_dc/" xmlns:dc="http://purl.org/dc/elements/1.1/" xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" xsi:schemaLocation="http://www.openarchives.org/OAI/2.0/oai_dc/ http://www.openarchives.org/OAI/2.0/oai_dc.xsd">''')
	try:
		current.write("<dc:title>"+f["Title"]+"</dc:title>")
	except:
		pass
	try:
		creators = f["Creators"].split(";")
		for creator in creators:
			current.write("<dc:creator>"+creator+"</dc:creator>")
	except:
		pass
	try:
		subjects = f["Subjects"].split(";")
		for subject in subjects:
			current.write("<dc:subject>"+subject+"</dc:subject>")
	except:
		pass
	try:
		contributors = f["Contributors"].split(";")
		for contributor in contributors:
			current.write("<dc:contributor>"+contributor+"</dc:contributor>")
	except:
		pass
	try:
		current.write("<dc:description>"+f["Description"]+"</dc:description>")
	except:
		pass
	try:
		current.write("<dc:date>"+f["Date"]+"</dc:date>")
	except:
		pass
	try:
		types = f["Type"].split(";")
		for type in types:
			current.write("<dc:type>"+type+"</dc:type>")
	except:
		pass
	try:
		current.write("<dc:format>"+f["Format"]+"</dc:format>")
	except:
		pass
	try:
		current.write("<dc:identifier>"+f["Identifier"]+"</dc:identifier>")
	except:
		pass
	try:
		current.write("<dc:source>"+f["Source"]+"</dc:source>")
	except:
		pass
	try:
		current.write("<dc:language>"+f["Language"]+"</dc:language>")
	except:
		pass
	try:
		current.write("<dc:relation>"+f["Relation"]+"</dc:relation>")
	except:
		pass
	try:
		current.write("<dc:coverage>"+f["Coverage"]+"</dc:coverage>")
	except:
		pass
	try:
		current.write("<dc:rights>"+f["Rights"]+"</dc:rights>")
	except:
		pass
	current.write("</oai_dc:dc>")
	print "In progress... completed one file..."
print "Completed"
